#!/usr/bin/env python3
"""
Runner for SpreadsheetBench benchmark.

Simplified implementation based on paper insights:
1. Baseline: Single-round code generation
2. ReAct: Multi-round with execution feedback (paper's key improvement)

Usage:
    python runner.py --limit 20 --mode baseline
    python runner.py --limit 20 --mode react --max-turns 5
"""

import argparse
import json
import os
import re
import shutil
import subprocess
import tempfile
from datetime import datetime

import anthropic

from data_loader import load_spreadsheetbench, load_sample_data
from evaluator import evaluate_instruction, calculate_metrics


client = anthropic.Anthropic()
DEFAULT_MODEL = "claude-sonnet-4-20250514"


def extract_code(response: str) -> str:
    """Extract Python code from model response."""
    patterns = [
        r"```python\n(.*?)```",
        r"```py\n(.*?)```",
        r"```\n(.*?)```",
    ]
    for pattern in patterns:
        matches = re.findall(pattern, response, re.DOTALL)
        if matches:
            return max(matches, key=len).strip()
    return response.strip()


def call_claude(messages: list, model: str = DEFAULT_MODEL, max_tokens: int = 4096) -> str:
    """Call Claude API."""
    response = client.messages.create(
        model=model,
        max_tokens=max_tokens,
        messages=messages,
    )
    return response.content[0].text


def execute_code(code: str, input_path: str, output_path: str, timeout: int = 30) -> dict:
    """Execute Python code and return result."""
    try:
        shutil.copy(input_path, output_path)
    except Exception as e:
        return {"success": False, "error": f"Copy failed: {e}"}

    wrapper = f'''
import sys
sys.path.insert(0, '.')
file_path = r"{output_path}"
input_file = r"{input_path}"
output_file = r"{output_path}"

{code}
'''
    try:
        with tempfile.NamedTemporaryFile(mode='w', suffix='.py', delete=False) as f:
            f.write(wrapper)
            wrapper_path = f.name

        proc = subprocess.run(
            ['python', wrapper_path],
            capture_output=True,
            text=True,
            timeout=timeout,
            cwd=os.path.dirname(input_path) or '.',
        )

        if proc.returncode != 0:
            error_lines = proc.stderr.strip().split('\n')
            return {"success": False, "error": '\n'.join(error_lines[-10:])}

        return {"success": True, "output": proc.stdout}

    except subprocess.TimeoutExpired:
        return {"success": False, "error": f"Timeout after {timeout}s"}
    except Exception as e:
        return {"success": False, "error": str(e)}
    finally:
        if 'wrapper_path' in locals():
            try:
                os.remove(wrapper_path)
            except:
                pass


# =============================================================================
# BASELINE: Single-round (aligned with official inference_single.py)
# =============================================================================

PROMPT_FORMAT_SINGLE = """You are a spreadsheet manipulation agent. I will provide you with six types of information:
1. Instruction: the instruction of the spreadsheet manipulation task
2. Spreadsheet Path: the path of the spreadsheet file to be manipulated
3. Spreadsheet Content: the content of the spreadsheet file (first rows of each sheet)
4. Instruction Type: the type of the instruction (Cell-Level or Sheet-Level Manipulation)
5. Answer Position: the cell range where the answer should be written
6. Output Path: the path of the output spreadsheet file

Instruction: {instruction}
Spreadsheet Path: {spreadsheet_path}
Spreadsheet Content:
{spreadsheet_content}
Instruction Type: {instruction_type}
Answer Position: {answer_position}
Output Path: {output_path}

Please generate Python code for the final solution. Use openpyxl library.
The code should:
1. Load the workbook from the spreadsheet path
2. Perform the required manipulation
3. Save the result to the output path

```python
from openpyxl import load_workbook

wb = load_workbook("{spreadsheet_path}")
# Your code here
wb.save("{output_path}")
```
"""


def generate_baseline(sample: dict, model: str = DEFAULT_MODEL) -> str:
    """Single-round code generation (aligned with official inference_single.py)."""
    prompt = PROMPT_FORMAT_SINGLE.format(
        instruction=sample["instruction"],
        spreadsheet_path="file_path",  # Use variable name
        spreadsheet_content=sample.get("preview", "")[:2000],
        instruction_type=sample["instruction_type"],
        answer_position=sample["answer_position"],
        output_path="file_path",  # Same as input for our setup
    )
    response = call_claude([{"role": "user", "content": prompt}], model=model)
    return extract_code(response)


# =============================================================================
# REACT: Multi-round with execution feedback (aligned with official script)
# https://github.com/RUCKBReasoning/SpreadsheetBench/blob/main/inference/inference_multiple.py
# =============================================================================

PROMPT_DF_RCT_FORMAT = """You are a spreadsheet manipulation agent. I will provide you with six types of information:
1. Instruction: the instruction of the spreadsheet manipulation task
2. Spreadsheet Path: the path of the spreadsheet file to be manipulated
3. Spreadsheet Content: the content of the spreadsheet file (first rows of each sheet)
4. Instruction Type: the type of the instruction (Cell-Level or Sheet-Level Manipulation)
5. Answer Position: the cell range where the answer should be written
6. Output Path: the path of the output spreadsheet file

Instruction: {instruction}
Spreadsheet Path: {spreadsheet_path}
Spreadsheet Content:
{spreadsheet_content}
Instruction Type: {instruction_type}
Answer Position: {answer_position}
Output Path: {output_path}

The solution can be generated through {max_turn_num} rounds of interaction.
In each round, you can choose one of the following two actions:
1. Generate Python code to explore the spreadsheet or test your solution
2. Generate the final Python code solution

After each code execution, I will provide you with the execution result.
If there are errors, please fix them in the next round.

Please generate Python code using openpyxl library:
```python
from openpyxl import load_workbook

wb = load_workbook("{spreadsheet_path}")
# Your code here
wb.save("{output_path}")
```
"""


def generate_react(
    sample: dict,
    model: str = DEFAULT_MODEL,
    max_turns: int = 5,
    test_input: str = None,
    test_output: str = None,
) -> tuple:
    """
    Multi-round code generation with execution feedback.
    Aligned with official inference_multiple.py (row_react_exec setting).
    """
    messages = []
    turns_used = 0

    # Initial prompt (aligned with PROMPT_DF_RCT_FORMAT)
    init_prompt = PROMPT_DF_RCT_FORMAT.format(
        instruction=sample["instruction"],
        spreadsheet_path="file_path",
        spreadsheet_content=sample.get("preview", "")[:2000],
        instruction_type=sample["instruction_type"],
        answer_position=sample["answer_position"],
        output_path="file_path",
        max_turn_num=max_turns,
    )
    messages.append({"role": "user", "content": init_prompt})

    final_code = None

    for turn in range(max_turns):
        turns_used += 1

        # Get response from model
        response = call_claude(messages, model=model)
        messages.append({"role": "assistant", "content": response})

        # Extract code
        code = extract_code(response)
        final_code = code

        # If no test files, return after first response
        if not test_input or not test_output:
            break

        # Execute code and get feedback (aligned with official approach)
        result = execute_code(code, test_input, test_output)

        # Check if output file exists - if so, we're done
        if os.path.exists(test_output) and result["success"]:
            break

        # Append execution result as feedback (matching official script pattern)
        if result["success"]:
            exec_result = result.get("output", "Code executed successfully.")
            if not exec_result.strip():
                exec_result = "Code executed successfully. No output."
        else:
            exec_result = f"Error occur when running code.\n{result['error']}"

        messages.append({"role": "user", "content": exec_result})

    return final_code, turns_used


# =============================================================================
# Main Runner
# =============================================================================

def run_benchmark(
    limit: int = None,
    model: str = DEFAULT_MODEL,
    mode: str = "baseline",
    max_turns: int = 5,
    use_sample: bool = False,
    data_dir: str = None,
    output_dir: str = None,
    instruction_types: list = None,
):
    """Run benchmark."""

    print("=" * 60)
    print("SpreadsheetBench Benchmark")
    print("=" * 60)
    print(f"Model: {model}")
    print(f"Mode: {mode}")
    if mode in ["react", "all"]:
        print(f"Max turns: {max_turns}")

    # Determine which modes to run
    if mode == "all":
        modes_to_run = ["baseline", "react"]
    else:
        modes_to_run = [mode]

    # Load data
    if use_sample:
        samples = load_sample_data()
        print("Using sample data (no evaluation)")
    else:
        samples = load_spreadsheetbench(
            data_dir=data_dir,
            limit=limit,
            instruction_types=instruction_types,
        )

    if not samples:
        print("No samples. Exiting.")
        return

    # Output directory
    if output_dir is None:
        output_dir = tempfile.mkdtemp(prefix="spreadsheetbench_")
    os.makedirs(output_dir, exist_ok=True)
    print(f"Output: {output_dir}")

    # Results storage per mode
    all_results = {m: [] for m in modes_to_run}
    total_turns = {m: 0 for m in modes_to_run}

    for i, sample in enumerate(samples):
        print(f"\n[{i+1}/{len(samples)}] ID: {sample['id']}")
        print(f"  Type: {sample['instruction_type']}")
        print(f"  Task: {sample['instruction'][:80]}...")

        sample_dir = os.path.join(output_dir, str(sample['id']))
        os.makedirs(sample_dir, exist_ok=True)

        # Get test file for execution feedback
        test_input = None
        test_output = None
        if sample['test_cases']:
            tc = sample['test_cases'][0]
            test_input = tc['input_file']
            test_output = os.path.join(sample_dir, "test_output.xlsx")

        for run_mode in modes_to_run:
            try:
                if run_mode == "baseline":
                    code = generate_baseline(sample, model=model)
                    turns = 1
                else:  # react
                    code, turns = generate_react(
                        sample, model=model, max_turns=max_turns,
                        test_input=test_input, test_output=test_output,
                    )

                total_turns[run_mode] += turns

                # Save code
                with open(os.path.join(sample_dir, f"{run_mode}_code.py"), "w") as f:
                    f.write(code)

                # Evaluate
                if sample['test_cases']:
                    eval_result = evaluate_instruction(
                        code=code,
                        test_cases=sample['test_cases'],
                        answer_position=sample['answer_position'],
                        output_dir=sample_dir,
                    )
                    eval_result["id"] = sample["id"]
                    eval_result["instruction_type"] = sample["instruction_type"]
                    eval_result["turns"] = turns
                    eval_result["code_length"] = len(code)
                    all_results[run_mode].append(eval_result)

                    status = "✓" if eval_result["hard_restriction"] == 1 else "✗"
                    print(f"  [{run_mode}] {status} Soft: {eval_result['soft_restriction']:.0%}, Hard: {eval_result['hard_restriction']}, Turns: {turns}")
                else:
                    print(f"  [{run_mode}] (No test cases)")

            except Exception as e:
                print(f"  [{run_mode}] ERROR: {e}")
                if sample['test_cases']:
                    all_results[run_mode].append({
                        "id": sample["id"],
                        "instruction_type": sample["instruction_type"],
                        "soft_restriction": 0.0,
                        "hard_restriction": 0,
                        "error": str(e),
                    })

    # Summary
    print("\n" + "=" * 60)
    print("RESULTS")
    print("=" * 60)

    metrics_by_mode = {}
    for run_mode in modes_to_run:
        results = all_results[run_mode]
        if results:
            metrics = calculate_metrics(results)
            metrics_by_mode[run_mode] = metrics
            print(f"\n[{run_mode.upper()}]")
            print(f"  Soft Restriction: {metrics['soft_restriction_avg']:.1%}")
            print(f"  Hard Restriction: {metrics['hard_restriction_avg']:.1%}")
            print(f"  Avg Turns: {total_turns[run_mode] / len(samples):.1f}")

            if metrics['by_type']:
                for t, d in metrics['by_type'].items():
                    print(f"    {t}: Soft={d['soft_restriction_avg']:.1%}, Hard={d['hard_restriction_avg']:.1%}")

    # Comparison if running both
    if len(modes_to_run) > 1 and "baseline" in metrics_by_mode and "react" in metrics_by_mode:
        print("\n" + "-" * 40)
        print("COMPARISON (react vs baseline)")
        print("-" * 40)
        base = metrics_by_mode["baseline"]
        react = metrics_by_mode["react"]
        soft_diff = react['soft_restriction_avg'] - base['soft_restriction_avg']
        hard_diff = react['hard_restriction_avg'] - base['hard_restriction_avg']
        print(f"  Soft: {soft_diff:+.1%}")
        print(f"  Hard: {hard_diff:+.1%}")

    # Save
    with open(os.path.join(output_dir, "results.json"), "w") as f:
        json.dump({
            "model": model,
            "mode": mode,
            "max_turns": max_turns,
            "timestamp": datetime.now().isoformat(),
            "results": all_results,
            "metrics": metrics_by_mode,
        }, f, indent=2, default=str)

    print(f"\nSaved to: {output_dir}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=None)
    parser.add_argument("--model", type=str, default=DEFAULT_MODEL)
    parser.add_argument("--mode", type=str, default="baseline", choices=["baseline", "react", "all"])
    parser.add_argument("--max-turns", type=int, default=5)
    parser.add_argument("--sample", action="store_true")
    parser.add_argument("--data-dir", type=str, default=None)
    parser.add_argument("--output-dir", type=str, default=None)
    parser.add_argument("--cell-level", action="store_true")
    parser.add_argument("--sheet-level", action="store_true")

    args = parser.parse_args()

    instruction_types = None
    if args.cell_level:
        instruction_types = ["Cell-Level Manipulation"]
    elif args.sheet_level:
        instruction_types = ["Sheet-Level Manipulation"]

    run_benchmark(
        limit=args.limit,
        model=args.model,
        mode=args.mode,
        max_turns=args.max_turns,
        use_sample=args.sample,
        data_dir=args.data_dir,
        output_dir=args.output_dir,
        instruction_types=instruction_types,
    )
