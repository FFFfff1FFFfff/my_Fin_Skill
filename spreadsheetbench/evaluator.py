#!/usr/bin/env python3
"""
Evaluator for SpreadsheetBench benchmark.
Aligned with official evaluation: https://github.com/RUCKBReasoning/SpreadsheetBench/blob/main/evaluation/evaluation.py

OJ-style evaluation:
- Each instruction has multiple test cases (usually 3)
- Soft Restriction: % of test cases that pass
- Hard Restriction: 1 if all pass, 0 otherwise
"""

import os
import re
from datetime import datetime, time
from typing import Optional, Union
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries


def transform_value(value) -> Union[str, float, None]:
    """
    Transform cell value for comparison (aligned with official script).
    - Numeric values rounded to 2 decimals
    - Datetime converted to Excel serial format
    - Time values truncated to milliseconds
    - String numbers parsed as floats when possible
    """
    if value is None:
        return None

    if value == "":
        return ""

    # Handle numbers - round to 2 decimals
    if isinstance(value, (int, float)):
        return round(float(value), 2)

    # Handle datetime - convert to Excel serial number
    if isinstance(value, datetime):
        # Excel serial date: days since 1899-12-30
        delta = value - datetime(1899, 12, 30)
        return round(delta.days + delta.seconds / 86400, 2)

    # Handle time - truncate to milliseconds and convert to string
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")

    # Handle strings
    if isinstance(value, str):
        value = value.strip()
        # Try to parse as float
        try:
            return round(float(value), 2)
        except ValueError:
            return value

    return str(value)


def compare_cell_value(v1, v2) -> bool:
    """
    Compare two cell values (aligned with official script).
    """
    v1 = transform_value(v1)
    v2 = transform_value(v2)

    # Handle empty/None equivalence
    if (v1 == "" and v2 is None) or (v1 is None and v2 == ""):
        return True
    if (v1 == "" and v2 == "") or (v1 is None and v2 is None):
        return True

    # Type mismatch
    if type(v1) != type(v2):
        return False

    return v1 == v2


def parse_answer_position(answer_position: str) -> tuple:
    """
    Parse answer position string to get sheet name and cell range.

    Formats:
    - "D2:D10" -> (None, "D2:D10")
    - "Sheet1!A1:Z100" -> ("Sheet1", "A1:Z100")
    - "Sheet1" -> ("Sheet1", None)  # entire sheet
    """
    if "!" in answer_position:
        parts = answer_position.split("!")
        sheet_name = parts[0].strip("'\"")
        cell_range = parts[1] if len(parts) > 1 else None
        return sheet_name, cell_range
    elif ":" in answer_position:
        return None, answer_position
    else:
        # Could be just a sheet name or a single cell
        if re.match(r"^[A-Z]+\d+$", answer_position):
            return None, answer_position
        return answer_position, None


def get_cell_range_values(ws, cell_range: Optional[str]) -> dict:
    """
    Extract raw values from a cell range in a worksheet.
    Values are NOT transformed here - transformation happens in compare_cell_value.

    Returns dict mapping (row, col) -> raw_value
    """
    values = {}

    if cell_range is None:
        # Get all cells with values
        for row in ws.iter_rows():
            for cell in row:
                values[(cell.row, cell.column)] = cell.value
    else:
        try:
            # Parse range like "A1:D10"
            min_col, min_row, max_col, max_row = range_boundaries(cell_range)

            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    values[(row, col)] = cell.value
        except Exception:
            # Try as single cell
            try:
                cell = ws[cell_range]
                if hasattr(cell, 'value'):
                    values[(cell.row, cell.column)] = cell.value
            except Exception:
                pass

    return values


def compare_cell_values(expected: dict, actual: dict) -> tuple:
    """
    Compare expected and actual cell values using official comparison logic.

    Returns (is_match, details)
    """
    all_keys = set(expected.keys()) | set(actual.keys())
    mismatches = []

    for key in all_keys:
        exp_val = expected.get(key)
        act_val = actual.get(key)

        # Use official compare_cell_value function
        if not compare_cell_value(exp_val, act_val):
            mismatches.append(f"Cell {key}: expected '{exp_val}', got '{act_val}'")

    is_match = len(mismatches) == 0
    return is_match, mismatches


def compare_workbooks(
    answer_path: str,
    output_path: str,
    answer_position: str,
    check_colors: bool = False,
) -> tuple:
    """
    Compare answer workbook with output workbook.

    Returns (is_match, details)
    """
    if not os.path.exists(answer_path):
        return False, ["Answer file not found"]

    if not os.path.exists(output_path):
        return False, ["Output file not found"]

    try:
        wb_answer = load_workbook(answer_path, data_only=True)
        wb_output = load_workbook(output_path, data_only=True)
    except Exception as e:
        return False, [f"Error loading workbooks: {e}"]

    # Parse answer position
    sheet_name, cell_range = parse_answer_position(answer_position)

    # Get target sheet
    if sheet_name:
        if sheet_name not in wb_answer.sheetnames:
            return False, [f"Sheet '{sheet_name}' not found in answer"]
        if sheet_name not in wb_output.sheetnames:
            return False, [f"Sheet '{sheet_name}' not found in output"]
        ws_answer = wb_answer[sheet_name]
        ws_output = wb_output[sheet_name]
    else:
        # Use active sheet or first sheet
        ws_answer = wb_answer.active
        ws_output = wb_output.active

    # Get values from answer position
    expected_values = get_cell_range_values(ws_answer, cell_range)
    actual_values = get_cell_range_values(ws_output, cell_range)

    # Compare values
    is_match, mismatches = compare_cell_values(expected_values, actual_values)

    # Close workbooks
    wb_answer.close()
    wb_output.close()

    return is_match, mismatches


def evaluate_single_test_case(
    code: str,
    input_path: str,
    answer_path: str,
    output_path: str,
    answer_position: str,
    timeout: int = 30,
) -> dict:
    """
    Evaluate a single test case.

    Returns dict with:
    - success: bool
    - error: str or None
    - match: bool (if output matches answer)
    - details: list of mismatch details
    """
    import shutil
    import subprocess
    import tempfile

    result = {
        "success": False,
        "error": None,
        "match": False,
        "details": [],
    }

    # Copy input file to output location
    try:
        shutil.copy(input_path, output_path)
    except Exception as e:
        result["error"] = f"Failed to copy input file: {e}"
        return result

    # Create a wrapper script that executes the code
    wrapper_code = f'''
import sys
sys.path.insert(0, '.')

# Set file paths as variables the code might expect
input_file = r"{input_path}"
output_file = r"{output_path}"
file_path = r"{output_path}"

# Execute the generated code
{code}
'''

    # Write wrapper to temp file and execute
    try:
        with tempfile.NamedTemporaryFile(mode='w', suffix='.py', delete=False) as f:
            f.write(wrapper_code)
            wrapper_path = f.name

        # Execute with timeout
        proc = subprocess.run(
            ['python', wrapper_path],
            capture_output=True,
            text=True,
            timeout=timeout,
            cwd=os.path.dirname(input_path) or '.',
        )

        if proc.returncode != 0:
            result["error"] = f"Code execution failed:\n{proc.stderr}"
            return result

        result["success"] = True

    except subprocess.TimeoutExpired:
        result["error"] = f"Code execution timed out after {timeout}s"
        return result
    except Exception as e:
        result["error"] = f"Execution error: {e}"
        return result
    finally:
        # Clean up temp file
        if 'wrapper_path' in locals():
            try:
                os.remove(wrapper_path)
            except:
                pass

    # Compare output with answer
    is_match, details = compare_workbooks(answer_path, output_path, answer_position)
    result["match"] = is_match
    result["details"] = details

    return result


def evaluate_instruction(
    code: str,
    test_cases: list,
    answer_position: str,
    output_dir: str,
) -> dict:
    """
    Evaluate code on all test cases for an instruction.

    Returns dict with:
    - test_results: list of individual test results
    - soft_restriction: float (% of test cases passed)
    - hard_restriction: int (1 if all passed, 0 otherwise)
    """
    results = {
        "test_results": [],
        "soft_restriction": 0.0,
        "hard_restriction": 0,
    }

    if not test_cases:
        return results

    passed_count = 0

    for tc in test_cases:
        test_num = tc["test_num"]
        input_path = tc["input_file"]
        answer_path = tc.get("answer_file")

        # Create output path
        output_filename = os.path.basename(input_path).replace("_input.xlsx", "_output.xlsx")
        output_path = os.path.join(output_dir, output_filename)

        if answer_path is None:
            test_result = {
                "test_num": test_num,
                "success": False,
                "error": "No answer file",
                "match": False,
            }
        else:
            test_result = evaluate_single_test_case(
                code=code,
                input_path=input_path,
                answer_path=answer_path,
                output_path=output_path,
                answer_position=answer_position,
            )
            test_result["test_num"] = test_num

        results["test_results"].append(test_result)

        if test_result.get("match"):
            passed_count += 1

    # Calculate metrics
    total = len(test_cases)
    results["soft_restriction"] = passed_count / total if total > 0 else 0.0
    results["hard_restriction"] = 1 if passed_count == total else 0

    return results


def calculate_metrics(all_results: list) -> dict:
    """
    Calculate aggregate metrics across all instructions.

    Returns dict with:
    - soft_restriction_avg: Average soft restriction score
    - hard_restriction_avg: Average hard restriction score (% of instructions fully correct)
    - total_instructions: Number of instructions evaluated
    - by_type: Breakdown by instruction type
    """
    if not all_results:
        return {
            "soft_restriction_avg": 0.0,
            "hard_restriction_avg": 0.0,
            "total_instructions": 0,
            "by_type": {},
        }

    soft_scores = []
    hard_scores = []
    by_type = {}

    for r in all_results:
        soft = r.get("soft_restriction", 0.0)
        hard = r.get("hard_restriction", 0)
        itype = r.get("instruction_type", "Unknown")

        soft_scores.append(soft)
        hard_scores.append(hard)

        if itype not in by_type:
            by_type[itype] = {"soft_scores": [], "hard_scores": []}
        by_type[itype]["soft_scores"].append(soft)
        by_type[itype]["hard_scores"].append(hard)

    # Calculate averages
    metrics = {
        "soft_restriction_avg": sum(soft_scores) / len(soft_scores),
        "hard_restriction_avg": sum(hard_scores) / len(hard_scores),
        "total_instructions": len(all_results),
        "by_type": {},
    }

    for itype, data in by_type.items():
        metrics["by_type"][itype] = {
            "soft_restriction_avg": sum(data["soft_scores"]) / len(data["soft_scores"]),
            "hard_restriction_avg": sum(data["hard_scores"]) / len(data["hard_scores"]),
            "count": len(data["soft_scores"]),
        }

    return metrics


if __name__ == "__main__":
    # Test evaluator functions
    print("Testing evaluator functions...")

    # Test parse_answer_position
    test_positions = [
        "D2:D10",
        "Sheet1!A1:Z100",
        "'My Sheet'!B2:C5",
        "Sheet1",
        "A1",
    ]

    print("\nParsing answer positions:")
    for pos in test_positions:
        sheet, cell_range = parse_answer_position(pos)
        print(f"  '{pos}' -> sheet='{sheet}', range='{cell_range}'")

    # Test normalize_value
    test_values = [
        123,
        123.456789,
        "  Hello World  ",
        None,
        datetime(2024, 1, 15, 10, 30),
    ]

    print("\nNormalizing values:")
    for val in test_values:
        normalized = normalize_value(val)
        print(f"  {val!r} -> {normalized!r}")
