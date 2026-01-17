#!/usr/bin/env python3
"""
Data loader for SpreadsheetBench benchmark.

SpreadsheetBench: Real-world spreadsheet manipulation tasks from Excel forums.
- 912 instructions with 2,729 test cases (avg 3 per instruction)
- Task: Generate Python code to manipulate Excel spreadsheets
- Evaluation: OJ-style (code must work on all test cases)

Data format:
{
    "id": "unique_id",
    "instruction": "The manipulation task description",
    "spreadsheet_path": "path/to/test_cases/",
    "instruction_type": "Cell-Level" | "Sheet-Level",
    "answer_position": "D2:D10" or "Sheet1!A1:Z100"
}
"""

import json
import os
import tarfile
import tempfile
import shutil
from typing import Optional
import pandas as pd
from openpyxl import load_workbook


# Data URLs
DATA_URLS = {
    "sample_200": "https://raw.githubusercontent.com/RUCKBReasoning/SpreadsheetBench/main/data/sample_data_200.tar.gz",
    "full_912": "https://raw.githubusercontent.com/RUCKBReasoning/SpreadsheetBench/main/data/spreadsheetbench_912_v0.1.tar.gz",
    "verified_400": "https://raw.githubusercontent.com/RUCKBReasoning/SpreadsheetBench/main/data/spreadsheetbench_verified_400.tar.gz",
}


def download_and_extract(url: str, extract_dir: str) -> str:
    """Download and extract a tar.gz file."""
    import urllib.request

    # Download to temp file
    print(f"Downloading from {url}...")
    tar_path = os.path.join(extract_dir, "data.tar.gz")
    urllib.request.urlretrieve(url, tar_path)

    # Extract
    print(f"Extracting to {extract_dir}...")
    with tarfile.open(tar_path, "r:gz") as tar:
        tar.extractall(extract_dir)

    # Remove tar file
    os.remove(tar_path)

    # Find the extracted directory (usually has a specific name)
    for item in os.listdir(extract_dir):
        item_path = os.path.join(extract_dir, item)
        if os.path.isdir(item_path):
            return item_path

    return extract_dir


def read_spreadsheet_preview(xlsx_path: str, max_rows: int = 10) -> str:
    """Read first few rows of a spreadsheet as string preview."""
    try:
        # Try reading with pandas for a clean preview
        xl = pd.ExcelFile(xlsx_path)
        previews = []

        for sheet_name in xl.sheet_names[:3]:  # Limit to first 3 sheets
            df = pd.read_excel(xl, sheet_name=sheet_name, header=None, nrows=max_rows)
            preview = f"=== Sheet: {sheet_name} ===\n"
            preview += df.to_string(index=False, header=False)
            previews.append(preview)

        return "\n\n".join(previews)
    except Exception as e:
        return f"Error reading spreadsheet: {e}"


def read_spreadsheet_schema(xlsx_path: str) -> dict:
    """Extract schema information from a spreadsheet."""
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        schema = {
            "sheets": [],
            "total_rows": 0,
            "total_cols": 0,
        }

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_info = {
                "name": sheet_name,
                "max_row": ws.max_row,
                "max_col": ws.max_column,
                "dimensions": ws.dimensions,
            }
            schema["sheets"].append(sheet_info)
            schema["total_rows"] = max(schema["total_rows"], ws.max_row or 0)
            schema["total_cols"] = max(schema["total_cols"], ws.max_column or 0)

        wb.close()
        return schema
    except Exception as e:
        return {"error": str(e)}


def load_spreadsheetbench(
    data_dir: Optional[str] = None,
    dataset_type: str = "sample_200",
    limit: Optional[int] = None,
    instruction_types: Optional[list] = None,
) -> list:
    """
    Load SpreadsheetBench dataset.

    Args:
        data_dir: Path to extracted data directory. If None, downloads automatically.
        dataset_type: One of "sample_200", "full_912", "verified_400"
        limit: Maximum number of samples to load
        instruction_types: Filter by instruction type ["Cell-Level", "Sheet-Level"]

    Returns:
        List of sample dictionaries with keys:
        - id: Unique identifier
        - instruction: The manipulation task
        - instruction_type: Cell-Level or Sheet-Level
        - answer_position: Target cells for output
        - spreadsheet_path: Path to test case folder
        - test_cases: List of test case info (input/answer file paths)
        - preview: String preview of first test case spreadsheet
        - schema: Schema information about the spreadsheet
    """
    # Handle data directory
    if data_dir is None:
        # Create temp directory and download
        data_dir = os.path.join(tempfile.gettempdir(), f"spreadsheetbench_{dataset_type}")
        if not os.path.exists(data_dir):
            os.makedirs(data_dir)
            url = DATA_URLS.get(dataset_type)
            if url:
                data_dir = download_and_extract(url, data_dir)
            else:
                raise ValueError(f"Unknown dataset type: {dataset_type}")

    # Find dataset.json
    dataset_json_path = None
    for root, dirs, files in os.walk(data_dir):
        if "dataset.json" in files:
            dataset_json_path = os.path.join(root, "dataset.json")
            data_dir = root
            break

    if dataset_json_path is None:
        raise FileNotFoundError(f"dataset.json not found in {data_dir}")

    # Load dataset metadata
    print(f"Loading dataset from {dataset_json_path}...")
    with open(dataset_json_path, 'r', encoding='utf-8') as f:
        dataset = json.load(f)

    samples = []

    for item in dataset:
        # Filter by instruction type
        if instruction_types and item.get("instruction_type") not in instruction_types:
            continue

        item_id = item["id"]
        spreadsheet_rel_path = item.get("spreadsheet_path", "")

        # Find test case files
        test_cases = []
        spreadsheet_full_path = os.path.join(data_dir, spreadsheet_rel_path)

        # Look for test cases (usually 3 per instruction)
        for test_num in [1, 2, 3]:
            input_file = os.path.join(spreadsheet_full_path, f"{test_num}_{item_id}_input.xlsx")
            answer_file = os.path.join(spreadsheet_full_path, f"{test_num}_{item_id}_answer.xlsx")

            if os.path.exists(input_file):
                test_cases.append({
                    "test_num": test_num,
                    "input_file": input_file,
                    "answer_file": answer_file if os.path.exists(answer_file) else None,
                })

        if not test_cases:
            # Try alternative naming patterns
            if os.path.isdir(spreadsheet_full_path):
                for fname in os.listdir(spreadsheet_full_path):
                    if fname.endswith("_input.xlsx"):
                        test_num = fname.split("_")[0]
                        input_file = os.path.join(spreadsheet_full_path, fname)
                        answer_fname = fname.replace("_input.xlsx", "_answer.xlsx")
                        answer_file = os.path.join(spreadsheet_full_path, answer_fname)
                        test_cases.append({
                            "test_num": test_num,
                            "input_file": input_file,
                            "answer_file": answer_file if os.path.exists(answer_file) else None,
                        })

        # Get preview and schema from first test case
        preview = ""
        schema = {}
        if test_cases:
            first_input = test_cases[0]["input_file"]
            preview = read_spreadsheet_preview(first_input)
            schema = read_spreadsheet_schema(first_input)

        sample = {
            "id": item_id,
            "instruction": item.get("instruction", ""),
            "instruction_type": item.get("instruction_type", ""),
            "answer_position": item.get("answer_position", ""),
            "spreadsheet_path": spreadsheet_full_path,
            "test_cases": test_cases,
            "preview": preview,
            "schema": schema,
        }

        samples.append(sample)

        if limit and len(samples) >= limit:
            break

    print(f"Loaded {len(samples)} samples")

    # Print distribution
    type_counts = {}
    for s in samples:
        t = s["instruction_type"]
        type_counts[t] = type_counts.get(t, 0) + 1

    print("\nInstruction Type Distribution:")
    for t, count in sorted(type_counts.items()):
        print(f"  {t}: {count}")

    return samples


def load_sample_data() -> list:
    """Load built-in sample data for testing without downloading."""
    return [
        {
            "id": "sample_1",
            "instruction": "Extract all values from column A that are greater than 100 and put them in column D",
            "instruction_type": "Cell-Level",
            "answer_position": "D2:D10",
            "spreadsheet_path": "",
            "test_cases": [],
            "preview": "=== Sheet: Sheet1 ===\n   A    B    C\n  50  ABC  100\n 150  DEF  200\n  75  GHI  300\n 200  JKL  400",
            "schema": {"sheets": [{"name": "Sheet1", "max_row": 5, "max_col": 3}]},
        },
        {
            "id": "sample_2",
            "instruction": "Highlight all cells in column B that contain the word 'error' with red background",
            "instruction_type": "Cell-Level",
            "answer_position": "B:B",
            "spreadsheet_path": "",
            "test_cases": [],
            "preview": "=== Sheet: Sheet1 ===\n   A           B\n   1  success\n   2  error found\n   3  completed\n   4  error: timeout",
            "schema": {"sheets": [{"name": "Sheet1", "max_row": 5, "max_col": 2}]},
        },
        {
            "id": "sample_3",
            "instruction": "Create a summary table showing the sum of column C grouped by the values in column A",
            "instruction_type": "Sheet-Level",
            "answer_position": "E1:F10",
            "spreadsheet_path": "",
            "test_cases": [],
            "preview": "=== Sheet: Sheet1 ===\n   A    B     C\n Cat    X   100\n Dog    Y   200\n Cat    Z   150\n Dog    W   300",
            "schema": {"sheets": [{"name": "Sheet1", "max_row": 5, "max_col": 3}]},
        },
    ]


if __name__ == "__main__":
    # Test with sample data
    print("Testing with sample data...")
    samples = load_sample_data()

    for s in samples:
        print(f"\n{'='*60}")
        print(f"ID: {s['id']}")
        print(f"Type: {s['instruction_type']}")
        print(f"Instruction: {s['instruction']}")
        print(f"Answer Position: {s['answer_position']}")
        print(f"\nPreview:\n{s['preview']}")
